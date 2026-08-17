import io
from datetime import datetime
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app import db
from app.models import Empresa, CFDI, DownloadRequest

reportes_bp = Blueprint('reportes', __name__)


@reportes_bp.route('/reportes')
@login_required
def index():
    empresas = Empresa.query.filter_by(user_id=current_user.id).all()
    return render_template('reportes/index.html', empresas=empresas)


@reportes_bp.route('/reportes/iva/<int:empresa_id>')
@login_required
def iva(empresa_id):
    empresa = Empresa.query.get_or_404(empresa_id)
    if empresa.user_id != current_user.id:
        flash('No tienes acceso.', 'error')
        return redirect(url_for('reportes.index'))

    month = request.args.get('month', datetime.utcnow().month, type=int)
    year = request.args.get('year', datetime.utcnow().year, type=int)

    from sqlalchemy import func, extract

    iva_cobrado = db.session.query(func.sum(CFDI.iva_trasladado))\
        .filter_by(empresa_id=empresa_id, tipo_comprobante='I', estado='vigente')\
        .filter(extract('month', CFDI.fecha_emision) == month)\
        .filter(extract('year', CFDI.fecha_emision) == year)\
        .scalar() or 0

    iva_acreditable = db.session.query(func.sum(CFDI.iva_trasladado))\
        .filter_by(empresa_id=empresa_id, tipo_comprobante='E', estado='vigente')\
        .filter(extract('month', CFDI.fecha_emision) == month)\
        .filter(extract('year', CFDI.fecha_emision) == year)\
        .scalar() or 0

    iva_pagar = iva_cobrado - iva_acreditable

    isr_total = db.session.query(func.sum(CFDI.isr_retenido))\
        .filter_by(empresa_id=empresa_id, estado='vigente')\
        .filter(extract('month', CFDI.fecha_emision) == month)\
        .filter(extract('year', CFDI.fecha_emision) == year)\
        .scalar() or 0

    cfdis_detalle = CFDI.query.filter_by(empresa_id=empresa_id, estado='vigente')\
        .filter(extract('month', CFDI.fecha_emision) == month)\
        .filter(extract('year', CFDI.fecha_emision) == year)\
        .order_by(CFDI.fecha_emision.desc()).all()

    return render_template('reportes/iva.html',
                         empresa=empresa,
                         month=month, year=year,
                         iva_cobrado=iva_cobrado,
                         iva_acreditable=iva_acreditable,
                         iva_pagar=iva_pagar,
                         isr_total=isr_total,
                         cfdis_detalle=cfdis_detalle)


@reportes_bp.route('/reportes/isr/<int:empresa_id>')
@login_required
def isr(empresa_id):
    empresa = Empresa.query.get_or_404(empresa_id)
    if empresa.user_id != current_user.id:
        flash('No tienes acceso.', 'error')
        return redirect(url_for('reportes.index'))

    month = request.args.get('month', datetime.utcnow().month, type=int)
    year = request.args.get('year', datetime.utcnow().year, type=int)

    from sqlalchemy import func, extract

    isr_retenido = db.session.query(func.sum(CFDI.isr_retenido))\
        .filter_by(empresa_id=empresa_id, estado='vigente')\
        .filter(extract('month', CFDI.fecha_emision) == month)\
        .filter(extract('year', CFDI.fecha_emision) == year)\
        .scalar() or 0

    iva_retenido_total = db.session.query(func.sum(CFDI.iva_retenido))\
        .filter_by(empresa_id=empresa_id, estado='vigente')\
        .filter(extract('month', CFDI.fecha_emision) == month)\
        .filter(extract('year', CFDI.fecha_emision) == year)\
        .scalar() or 0

    cfdis_detalle = CFDI.query.filter(
        CFDI.empresa_id == empresa_id,
        CFDI.estado == 'vigente',
        (CFDI.isr_retenido > 0) | (CFDI.iva_retenido > 0),
        extract('month', CFDI.fecha_emision) == month,
        extract('year', CFDI.fecha_emision) == year,
    ).order_by(CFDI.fecha_emision.desc()).all()

    return render_template('reportes/isr.html',
                         empresa=empresa,
                         month=month, year=year,
                         isr_retenido=isr_retenido,
                         iva_retenido_total=iva_retenido_total,
                         cfdis_detalle=cfdis_detalle)


@reportes_bp.route('/reportes/exportar-excel/<int:empresa_id>')
@login_required
def exportar_excel(empresa_id):
    empresa = Empresa.query.get_or_404(empresa_id)
    if empresa.user_id != current_user.id:
        flash('No tienes acceso.', 'error')
        return redirect(url_for('reportes.index'))

    tipo = request.args.get('tipo', '')
    estado = request.args.get('estado', '')
    month = request.args.get('month', '', type=str)
    year = request.args.get('year', '', type=str)

    query = CFDI.query.filter_by(empresa_id=empresa.id)
    if tipo:
        query = query.filter_by(tipo_comprobante=tipo)
    if estado:
        query = query.filter_by(estado=estado)

    from sqlalchemy import extract
    if month:
        query = query.filter(extract('month', CFDI.fecha_emision) == int(month))
    if year:
        query = query.filter(extract('year', CFDI.fecha_emision) == int(year))

    cfdis = query.order_by(CFDI.fecha_emision.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'CFDIS'

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
    money_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'UUID', 'Tipo', 'Fecha Emision', 'Fecha Timbrado', 'Serie', 'Folio',
        'RFC Emisor', 'Nombre Emisor', 'RFC Receptor', 'Nombre Receptor',
        'Metodo Pago', 'Forma Pago', 'Moneda', 'Tipo Cambio',
        'Subtotal', 'IVA Trasladado', 'ISR Retenido', 'IVA Retenido',
        'Total Impuestos', 'Total', 'Estado', 'Uso CFDI',
        'Clave Prod/Serv', 'Cantidad', 'Descripcion Concepto'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    tipo_nombres = {'I': 'Ingreso', 'E': 'Egreso', 'T': 'Traslado', 'N': 'Nomina', 'P': 'Pago', 'R': 'Retencion'}
    metodo_nombres = {'PUE': 'Pago en una sola exhibicion', 'PPD': 'Pago en parcialidades'}
    forma_nombres = {
        '01': 'Efectivo', '02': 'Cheque nominativo', '03': 'Transferencia',
        '04': 'Tarjeta de credito', '05': 'Monedero electronico',
        '06': 'Dinero electronico', '08': 'Vales de despensa',
        '12': 'Dacion en pago', '13': 'Pago por servicios',
        '14': 'Anticipo', '15': 'Pago por cuenta ajena',
    }

    import json as _json
    current_row = 2
    for cf in cfdis:
        conceptos = []
        if cf.conceptos_json:
            try:
                conceptos = _json.loads(cf.conceptos_json)
            except Exception:
                pass
        if not conceptos:
            conceptos = [{'claveProdServ': '', 'cantidad': 0, 'descripcion': ''}]

        for concepto in conceptos:
            data = [
                cf.uuid,
                tipo_nombres.get(cf.tipo_comprobante, cf.tipo_comprobante),
                cf.fecha_emision.strftime('%Y-%m-%d %H:%M') if cf.fecha_emision else '',
                cf.fecha_timbrado.strftime('%Y-%m-%d %H:%M') if cf.fecha_timbrado else '',
                cf.serie or '',
                cf.folio or '',
                cf.rfc_emisor,
                cf.nombre_emisor,
                cf.rfc_receptor,
                cf.nombre_receptor,
                metodo_nombres.get(cf.metodo_pago, cf.metodo_pago or ''),
                forma_nombres.get(cf.forma_pago, cf.forma_pago or ''),
                cf.moneda,
                cf.tipo_cambio,
                cf.subtotal,
                cf.iva_trasladado,
                cf.isr_retenido,
                cf.iva_retenido,
                cf.impuestos,
                cf.total,
                cf.estado.capitalize(),
                cf.uso_cfdi,
                concepto.get('claveProdServ', ''),
                concepto.get('cantidad', ''),
                concepto.get('descripcion', ''),
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
                if col in (15, 16, 17, 18, 19, 20):
                    cell.number_format = '#,##0.00'
                    cell.fill = money_fill
                if col in (14,):
                    cell.number_format = '#,##0.000000'
            current_row += 1

    total_row = current_row
    money_cols = {15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T'}
    for col_num, col_letter in money_cols.items():
        cell = ws.cell(row=total_row, column=col_num, value=f'=SUM({col_letter}2:{col_letter}{total_row-1})')
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.border = thin_border

    ws.auto_filter.ref = f'A1:Y{current_row-1}'

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename = f'CFDIS_{empresa.rfc}_{timestamp}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
